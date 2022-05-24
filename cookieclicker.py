#Importing needed stuff
import tkinter as tk
from tkinter import messagebox
from random import randint as random
from os import getlogin
from time import time, sleep
import openpyxl
import threading

#Main window declaration and configuration
gameWindow = tk.Tk()
gameWindow.title("Coopy Clicker")
gameWindow.geometry("500x430")
gameWindow.resizable(False, False)

startTime = time()
saveTime = time()

gold = 0

saveInterval = 30

"------------------------------------FUNCTION------------------------------------"
def activateCheatCode():
    global cheatCode, gold, cheatCodeUsage
    if (cheatCode.get() == "iamrich"):
        gold = 1000000
        cheatCodeUsage += 1
        messagebox.showinfo("Succès", "Code de triche activé")
    else:
        messagebox.showerror("Erreur", "Code inconnu")

valuelist = [2, 10, 30, 45, 60, 120]

def setSaveInterval(value):
    global saveIntervalSlider, valuelist, saveInterval
    saveInterval = min(valuelist, key=lambda x:abs(x-int(value)))
    saveIntervalSlider.set(saveInterval)

# Expand the side menu, increasing the width by 10 every 5ms, once fully expanded set expanded to true
def expand():
    global cur_width, expanded
    cur_width += 10
    rep = gameWindow.after(5, expand)  # Repeat every 5ms
    sideMenu.config(width=cur_width)
    if cur_width >= max_w:
        expanded = True
        gameWindow.after_cancel(rep)  # Stop repeating
        fill()


# Close the side menu, decreasing the width by 10 every 5ms, once fully closed set expanded to false
def contract():
    global cur_width, expanded
    cur_width -= 10
    rep = gameWindow.after(5, contract)  # Repeat every 5ms
    sideMenu.config(width=cur_width)
    if cur_width <= min_w:
        expanded = False
        gameWindow.after_cancel(rep)  # Stop repeating
        fill()

def fill():
    if expanded:
        # Show a text and remove the image
        homeButton.config(text="Accueil", image="", width=14)
        shopButton.config(text="Améliorations", image="", width=14)
        settingsButton.config(text="Paramètres", image="", width=14)
        credits.config(text=f"Coopy Clicker\n\nDéveloppé par\nBrayan\nAlex\nYassine")

    else:
        # Bring the image back
        homeButton.config(image=homeImage, width=25)
        shopButton.config(image=shopImage, width=25)
        settingsButton.config(image=settingsImage, width=25)
        credits.config(text=f"")

def showCredits():
    messagebox.showinfo("Crédits et Remerciements", "Ressouces :\nIcônes : icons8.com\nCookie principal : Google Images\nQuelques bouts de code : stackoverflow.com\n\nRemerciements :\numi : Support moral <3")

"""
Depending on if you have enough gold or not the button will change to red or green when hovered, seems to only work on Windows tho so I did not implement it
Also I'm sure there is a better way to do it.
"""

def enterGreen(item):
    item.widget['background'] = "green"

def enterRed(item):
    item.widget['background'] = "red"

# Using #D4D6D6 color code instead of SystemButtonFace as it is Windows only
def leave(item):
    item.widget['background'] = "#D4D6D6"

"""
Theses functions handle frames switching
Using grid_remove() instead of grid_forget() allows us to keep all the settings, so using grid() will show it back without the need of specifying all the options again
"""

def showMainFrame():
    global mainFrame, itemShop, infosFrame, settingsFrame
    mainFrame.grid()
    infosFrame.grid()
    itemShop.grid_remove()
    settingsFrame.grid_remove()

def showShopFrame():
    global mainFrame, infosFrame, settingsFrame, moreGold, itemShop, goldsForGoldUpgrade, moreAutoClicker, goldsForAutoClicker, moreGoldInfoLabel, autoClickerAmount, moreAutoclickerInfoLabel, gold

    mainFrame.grid_remove()
    infosFrame.grid_remove()
    settingsFrame.grid_remove()
    itemShop.grid()

    moreGoldUpgrade = tk.Frame(itemShop, borderwidth=2, relief=tk.GROOVE)
    moreGoldUpgrade.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)

    moreAutoclickerUpgrade = tk.Frame(itemShop, borderwidth=2, relief=tk.GROOVE)
    moreAutoclickerUpgrade.grid(column=1, row=1, sticky=tk.E, padx=40, pady=5)

    moreGoldInfoLabel = tk.Label(moreGoldUpgrade,
                                 text=f"Augemente l'or par clic\nActuel : {format_number(goldUpgradeNumber)} or par clic\nCoûte : {format_number(goldsForGoldUpgrade)} or")
    moreGoldInfoLabel.grid(column=1, row=0, sticky=tk.W, padx=5, pady=5)
    moreGold = tk.Button(moreGoldUpgrade, text=f"Améliorer", width=15)
    moreGold.grid(column=1, row=1, sticky=tk.S, padx=5, pady=5)
    moreGold.bind("<ButtonRelease-1>", upgradeGold)
    moreGold.bind("<Leave>", leave)

    moreAutoclickerInfoLabel = tk.Label(moreAutoclickerUpgrade,
                                        text=f"Un ouvrier qui génère 1 d'or/s\nActuel : {format_number(autoClickerAmount)} or/s\nCoûte : {format_number(goldsForAutoClicker)} or")
    moreAutoclickerInfoLabel.grid(column=1, row=0, sticky=tk.W, padx=5, pady=5)
    moreAutoClicker = tk.Button(moreAutoclickerUpgrade, text=f"Améliorer", width=20)
    moreAutoClicker.grid(column=1, row=1, sticky=tk.W, padx=5, pady=5)
    moreAutoClicker.bind("<ButtonRelease-1>", upgradeAutoClicker)
    moreAutoClicker.bind("<Leave>", leave)

def showSettingsFrame():
    global mainFrame, itemShop, infosFrame, settingsFrame
    settingsFrame.grid()
    mainFrame.grid_remove()
    infosFrame.grid_remove()
    itemShop.grid_remove()


"""
End of the frame handling functions 
"""

# This is used to add gold, you have 50% chance to get between 1 and the upgrade number(default is 5), if you're unlucky you get 0
def incr(event):
    global totalCookies, gold, goldUpgradeNumber, randomAmountOfGold
    luck = random(0, 100)
    randomAmountOfGold = random(1, goldUpgradeNumber)
    if (luck > 50):
        gold += randomAmountOfGold
    totalCookies += 1
    counterLabel.config(text=f"{format_number(totalCookies)}")
    goldLabel.config(text=f"{format_number(gold)}")

def upgradeAutoClicker(event):
    global goldUpgradeNumber, moreGold, gold, goldsForAutoClicker, autoClickerAmount, moreAutoclickerInfoLabel
    if (gold >= goldsForAutoClicker):
        gold -= goldsForAutoClicker
        goldsForAutoClicker *= 1.15
        goldsForAutoClicker = int(goldsForAutoClicker)
        autoClickerAmount += 1
    moreAutoclickerInfoLabel.config(
        text=f"Un ouvrier qui génère 1 d'or/s\nActuel : {format_number(autoClickerAmount)} or/s\nCoûte : {format_number(goldsForAutoClicker)} or")
    autoclickerLabel.config(text=f"{format_number(autoClickerAmount)}")

def autoClicker(event):
    global totalCookies, goldUpgradeNumber, moreGold, gold, goldsForAutoClicker, autoClickerAmount, startTime, shopGoldLabel, autoclickerLabel
    if ((time() - startTime) > 1):
        startTime = time()
        print(f"{gold}")
        totalCookies += autoClickerAmount
        gold += autoClickerAmount
        counterLabel.config(text=f"{format_number(totalCookies)}")
        goldLabel.config(text=f"{format_number(gold)}")
        shopGoldLabel.config(text=f"{format_number(gold)}")
        autoclickerLabel.config(text=f"{format_number(autoClickerAmount)}")
        

# This upgrades the max amount of gold that can be generated
def upgradeGold(event):
    global goldUpgradeNumber, moreGold, gold, goldsForGoldUpgrade, moreGoldInfoLabel
    if (gold >= goldsForGoldUpgrade):
        gold -= goldsForGoldUpgrade
        goldUpgradeNumber *= 2
        goldsForGoldUpgrade *= 2.5
        goldsForGoldUpgrade = int(goldsForGoldUpgrade)
    moreGoldInfoLabel.config(
        text=f"Augemente l'or par clic\nActuel : {format_number(goldUpgradeNumber)} or par clic\nCoûte : {format_number(goldsForGoldUpgrade)} or")
    counterLabel.config(text=f"{format_number(totalCookies)}")
    goldLabel.config(text=f"{format_number(gold)}")

#This formats the thousands numbers to letters and also rounds them to 2 numbers after the comma
def format_number(num):
    global formatNumbers
    sign = ''
    if(num < 1000 or formatNumbers.get() == 0):
        num = num
    else:
        metric = {'T': 1000000000000, 'B': 1000000000, 'M': 1000000, 'K': 1000, '': 1}
        for i in metric:
            num_check = num / metric[i]

            if(num_check >= 1):
                num = num_check
                sign = i
                break
        num = round(num, 2)
    return f"{str(num).strip('.')}{sign}"
    
# This creates the shop window and all its UI
def openShop():
    showShopFrame()

# This updates the cookies counter on the main window

def setCounter(args):
    counterLabel.config(text=f"{format_number(totalCookies)}")

def loadFeature():
    global gold, goldUpgradeNumber,autoClickerAmount,totalCookies, goldsForAutoClicker, goldsForGoldUpgrade, cheatCodeUsage, counterLabel, saveInterval, formatNumbers
    workbook = openpyxl.load_workbook('save.xlsx')
    ws = workbook.active
    gold = ws['A2'].value
    totalCookies = ws['B2'].value
    goldUpgradeNumber = ws['C2'].value
    autoClickerAmount = ws['D2'].value
    saveInterval = ws['E2'].value
    goldsForAutoClicker = ws['F2'].value
    goldsForGoldUpgrade = ws['G2'].value
    formatNumbers.set(ws['H2'].value)
    cheatCodeUsage = ws['J2'].value
    print("loaded")
    counterLabel.config(text=f"{totalCookies}")

    
def saveFeature():
    global gold, goldUpgradeNumber,autoClickerAmount,totalCookies, goldsForAutoClicker, goldsForGoldUpgrade, cheatCodeUsage, saveInterval,saveTime, formatNumbers
    if((time() - saveTime) > saveInterval):
        saveTime = time()
        try:
            workbook = openpyxl.load_workbook('save.xlsx')
            ws = workbook.active
            ws['A1'] = "Gold"
            ws['A2'] = gold
            ws['B1'] = "Cookies"
            ws['B2'] = totalCookies
            ws['C1'] = "Click power"
            ws['C2'] = goldUpgradeNumber
            ws['D1'] = "Amount of Autoclickers"
            ws['D2'] = autoClickerAmount
            ws['E1'] = "Save interval"
            ws['E2'] = saveInterval
            ws['F1'] = "Autoclicker Price"
            ws['F2'] = goldsForAutoClicker
            ws['G1'] = "GoldUpgrade Price"
            ws['G2'] = goldsForGoldUpgrade
            ws['H1'] = "Format Numbers"
            ws['H2'] = formatNumbers.get()
            ws['J1'] = "Amount of cheat codes used"
            ws['J2'] = cheatCodeUsage
            workbook.save('save.xlsx')
            print("saved")
        except:
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws['A1'] = "Gold"
            ws['A2'] = 0
            ws['B1'] = "Cookies"
            ws['B2'] = 0
            ws['C1'] = "Click power"
            ws['C2'] = 5
            ws['D1'] = "Amount of autoclickers"
            ws['D2'] = 0
            ws['E1'] = "Save interval"
            ws['E2'] = 30
            ws['F1'] = "Autoclicker Price"
            ws['F2'] = 50
            ws['G1'] = "GoldUpgrade Price"
            ws['G2'] = 10
            ws['H1'] = "Format Numbers"
            ws['H2'] = 1
            ws['J1'] = "Amount of cheat codes used"
            ws['J2'] = 0
            workbook.save(filename = 'save.xlsx')
            print("created")



"------------------------------------VARIABLES------------------------------------"
#Declaring main logic variables
mainFrame = tk.Frame(gameWindow, borderwidth=2, relief=tk.GROOVE)
mainFrame.grid(column=1, row=0, sticky=tk.N, padx=5)

itemShop = tk.Frame(gameWindow, borderwidth=2, relief=tk.GROOVE)
itemShop.grid(column=1, row=0, sticky=tk.W, padx=5)
itemShop.grid_remove()

infosFrame = tk.Frame(gameWindow, borderwidth=2, relief=tk.GROOVE)
infosFrame.grid(column=1, row=0, sticky=tk.S, padx=5, pady=30)

settingsFrame = tk.Frame(gameWindow, borderwidth=2, relief=tk.GROOVE)
settingsFrame.grid(column=1, row=0, sticky=tk.W, padx=5, pady=0)
settingsFrame.grid_remove()

#Saves
savingFrame = tk.Frame(settingsFrame, borderwidth=2, relief=tk.GROOVE)
savingFrame.grid(column=0, row=0, sticky=tk.W, pady=10)
saveIntervalSlider = tk.Scale(savingFrame, from_=min(valuelist), to=max(valuelist), command=setSaveInterval, orient=tk.HORIZONTAL)
saveIntervalSlider.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
saveIntervalSlider.set(saveInterval)

#Global settings
globalSettingsFrame = tk.Frame(settingsFrame, borderwidth=2, relief=tk.GROOVE)
globalSettingsFrame.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
formatNumbers = tk.IntVar()
formatNumbers.set(1)
formatNumbersCB = tk.Checkbutton(globalSettingsFrame, text="Formatter les nombres", onvalue=1, offvalue=0, variable=formatNumbers)
formatNumbersCB.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)

#Cheat Code Category
cheatFrame = tk.Frame(settingsFrame, borderwidth=2, relief=tk.GROOVE)
cheatFrame.grid(column=0, row=2, sticky=tk.W, pady=10)
cheatLabel = tk.Label(cheatFrame, text=f"Activer code de triche")
cheatLabel.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
cheatCode = tk.StringVar()
cheatEntry = tk.Entry(cheatFrame, textvariable=cheatCode)
cheatEntry.grid(column=1, row=0, sticky=tk.W, padx=5, pady=5)
cheatButton = tk.Button(cheatFrame, text=f"Activer", command=activateCheatCode)
cheatButton.grid(column=1, row=1, sticky=tk.W, padx=5, pady=5)

#Shop display
shopGoldLabel = tk.Label(itemShop, text=f"Or : {format_number(gold)}")
shopGoldLabel.grid(column=1, row=0, sticky=tk.W, padx=5, pady=5)

# Define the icons to be shown
homeImage = tk.PhotoImage(file="assets/homeIcon.png")
shopImage = tk.PhotoImage(file="assets/shopIcon.png")
settingsImage = tk.PhotoImage(file="assets/settingsIcon.png")
goldImage = tk.PhotoImage(file="assets/goldIcon.png")
worker1Image = tk.PhotoImage(file="assets/workerIcon.png")
cookie1Image = tk.PhotoImage(file="assets/cookieIcon.png")

gameWindow.update() #for the width to get updated

#Theses are all the variables for the main UI
sideMenu = tk.Frame(gameWindow,bg="orange",width=35,height=gameWindow.winfo_height())
sideMenu.grid(column=0, row=0)

credits = tk.Label(sideMenu, text=f"", bg="orange")
credits.grid(column=0, row=3, padx=10, pady=200)

counterLabel = tk.Label(infosFrame, text="0")
counterLabel.grid(column=1, row=2, sticky=tk.W, padx=30, pady=5)
cookieImageLabel = tk.Label(infosFrame, image=cookie1Image).grid(column=1, row=2, sticky=tk.W, padx=5)

goldLabel = tk.Label(infosFrame, text="0")
goldLabel.grid(column=1, row=3, sticky=tk.W, padx=30, pady=5)
goldImageLabel = tk.Label(infosFrame, image=goldImage).grid(column=1, row=3, sticky=tk.W, padx=5)

autoclickerLabel = tk.Label(infosFrame, text="0")
autoclickerLabel.grid(column=1, row=4, sticky=tk.W, padx=30, pady=5)
autoclickerImageLabel = tk.Label(infosFrame, image=worker1Image).grid(column=1, row=4, sticky=tk.W, padx=5)

photo = tk.PhotoImage(file="assets/cookie.png")

incrButton = tk.Button(mainFrame, image=photo, relief=tk.FLAT, compound=tk.TOP, width=256, height=266)
incrButton.grid(column=1, row=1, sticky=tk.W, padx=15, pady=5)
incrButton.bind("<ButtonRelease-1>", incr)

# Make the buttons with the icons to be shown
homeButton = tk.Button(sideMenu, image=homeImage, bg="orange", relief=tk.FLAT, command=showMainFrame)
homeButton.grid(column=0, row=0, pady=10)

shopButton = tk.Button(sideMenu, image=shopImage, bg="orange", relief=tk.FLAT, command=openShop)
shopButton.grid(column=0, row=1)

settingsButton = tk.Button(sideMenu, image=settingsImage, bg="orange", relief=tk.FLAT, command=showSettingsFrame)
settingsButton.grid(column=0, row=2, pady=10)

saveLabel = tk.Label(savingFrame, text="Intervale de sauvegarde(en secondes)")
saveLabel.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)

warningLabel = tk.Label(savingFrame, text="La valeur minimal (2s) n'est PAS RECOMMANDÉE!!!")
warningLabel.grid(column=0, row=2, sticky=tk.W, padx=5, pady=5)

creditsButton = tk.Button(settingsFrame, text="Remerciements/Credits", command=showCredits)
creditsButton.grid(column=0, row=5, sticky=tk.W, padx=5, pady=5)

"""
Dude stop with that button already lmafo
loadButton = tk.Button(settingsFrame, text="Charger la sauvegarde!", command=loadFeature)
loadButton.grid(column=0, row=3, pady=10)
"""
"------------------------------------CODE------------------------------------"

totalCookies = 0

randomAmountOfGold = 0
moreGold = None
moreGoldInfoLabel = None
moreAutoclickerInfoLabel = None
moreAutoClicker = None
goldUpgradeNumber = 5
cheatCodeUsage = 0

#Declaring upgrade prices variables
goldsForGoldUpgrade = 10
goldsForAutoClicker = 50

autoClickerAmount = 0

"""
Begin side menu
"""
#Declaring variables for side menu
min_w = 35
max_w = 100
cur_width = min_w # Increasing width of the frame
expanded = False

# Bind to the frame, if entered or left
sideMenu.bind('<Enter>',lambda e: expand())
sideMenu.bind('<Leave>',lambda e: contract())

# So that it does not depend on the widgets inside the frame
sideMenu.grid_propagate(False)

"""
End side menu
"""

try:
    loadFeature()
except Exception as e:
    print("Erreur lors du chargement de la sauvegarde: %s" % e)

#changed from gameWindow.gameloop to this because allows to make our own game loop
while 1:
    autoClicker(gameWindow)
    saveFeature()
    gameWindow.update_idletasks()
    gameWindow.update()
    sleep(0.01)
